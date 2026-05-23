from openpyxl import load_workbook


class ExcelReader:

    def __init__(self, file_path):

        self.workbook = load_workbook(file_path)

        self.sheet = self.workbook["Sheet1"]

    def get_data(self):

        data = {}

        headers = []

        for cell in self.sheet[1]:
            headers.append(cell.value)

        for row in self.sheet.iter_rows(
                min_row=2,
                max_row=2,
                values_only=True):

            for key, value in zip(headers, row):

                data[key] = value

        return data